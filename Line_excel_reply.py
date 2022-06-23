import requests
"""
作品名稱     ： Line回文機器人
__author__ ： Ning 甯詠城
"""
# 一定要改的地方～～
auth_token = "hV0gN0T6RUKCAskSyGf4LbF1RxYUSqnHQqGZ9VEX4aV+TuyAnP4jYPGtOa+8L3Fvm4Eg0wlBXMj4nEwxWI68smcHAK/gVBOp3EQCsq4TzTiXIS1aZp/nwtmeSRjDHC59ngunTrdMGfR0PtUoZix06gdB04t89/1O/w1cDnyilFU="
import xlrd
import xlwt
from sys import version as python_version
from cgi import parse_header, parse_multipart
import socketserver as socketserver
import http.server
from http.server import SimpleHTTPRequestHandler as RequestHandler
from urllib.parse import parse_qs
import json
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import time                       # 時間
import lib_line_reply

###工作欄
wb = load_workbook("line.xlsx")  # '新竹縣美食資料.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheet = wb.active  # 打開一個工作欄
wb = load_workbook('line.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheetQA = wb["問答題"]         # 打開一個工作欄


###紀錄欄
wb_record = Workbook()                   # 初始化
sheet_record = wb_record.active                 # 新增一個工作欄
now = time.strftime("%x")        # 取得 現在的時間
a=["UserID"]
b=["text"]





"""def Line_quation(text): #判斷是否有值，沒有救回傳我不知道
    reply = lib_line_reply.openpyxl_GetRow(sheetQA, text, key_col=1, answer_col=2)
    if(reply!=""):
        reply_Text =Line_reply_Text(reply)
    else:
        reply_Text =Line_reply_Text("查無資料\n請輸入Hololive引導\n或者請撥電話xxxxxx")
    return reply_Text"""



def Line_quation(text):  #處理type   將他們分類成文字或者圖片
    reply = lib_line_reply.openpyxl_GetRow(sheetQA, text, key_col=1, answer_col=2)
    reply2 = lib_line_reply.openpyxl_GetRow(sheetQA, text, key_col=1, answer_col=3)
    if (reply2 == "text"):
        reply_Text = Line_reply_Text(reply)
        return reply_Text
    elif(reply2=="img"):
        reply_img = Line_reply_img(reply)
        return reply_img
    elif (reply2 == "template"):
        reply_template = Line_reply_template(reply)
        return reply_template
    else:
        reply_Text = Line_reply_Text("查無資料\n請輸入Hololive引導\n或者請撥電話xxxxxx")
        return reply_Text

def Line_reply_Text(reply_Text):   #reply type為文字檔的部分
    reply_Text ={
            "type": "text",  # 文字
            "text":reply_Text
        }
    return reply_Text

def Line_reply_img(reply_img):   #reply type為圖片檔的部分
    reply_img = {
            "type": "image",  # 圖片
            "originalContentUrl": reply_img,
            "previewImageUrl":reply_img
        }
    return reply_img
def Line_reply_template(reply_template):    #印出template
    split=reply_template.split(',',2)       #切割出以","為分割點的reply_template
    圖片=split[0]
    YT = split[1]
    TWITTER = split[2]


    reply_template={
        "type": "template",
        "altText": "This is a buttons template",
        "template": {
            "type": "buttons",
            "thumbnailImageUrl":圖片,
            "imageAspectRatio": "rectangle",
            "imageSize": "cover",
            "imageBackgroundColor": "#FFFFFF",
            "title": "Menu",
            "text": "Please select",
            "actions": [
                {
                    "type": "uri",
                    "label": "Youtube",
                    "uri": YT
                },
                {
                    "type": "uri",
                    "label": "Twitter",
                    "uri": TWITTER
                }
            ]
        }
    }
    return reply_template



class MyHandler(RequestHandler):
    def do_POST(self):
        varLen = int(self.headers['Content-Length'])  # 取得讀取進來的網路資料長度
        if varLen > 0:
            post_data = self.rfile.read(varLen)  # 讀取傳過來的資料
            data = json.loads(post_data)  # 把字串 轉成JSON
            print(data)
            replyToken = data['events'][0]['replyToken']  # 回傳要用Token
            userId = data['events'][0]['source']['userId']  # 傳資料過來的使用者是誰
            text = data['events'][0]['message']['text']  # 用戶的傳遞過來的文字內容
            傳過來的資料型態 = data['events'][0]['message']['type']  # 傳過來的資料型態

            a.append(data['events'][0]['source']['userId']) #用於紀錄ID
            b.append(data['events'][0]['message']['text'])  # 用於紀錄資料


        reply=Line_quation(text)   #文字
        # 請參考
        # https://developers.line.biz/zh-hant/docs/messaging-api/sending-messages/#methods-of-sending-message
        message = {

            "replyToken": replyToken,
            "messages": [
                reply

            ]
        }


        # 資料回傳 到 Line 的 https 伺服器
        hed = {'Authorization': 'Bearer ' + auth_token}
        url = 'https://api.line.me/v2/bot/message/reply'
        self.send_response(200)
        self.end_headers()
        requests.post(url, json=message, headers=hed)  # 把資料HTTP POST送出去


socketserver.TCPServer.allow_reuse_address = True  # 可以重複使用IP
httpd = socketserver.TCPServer(('0.0.0.0', 8888), MyHandler)  # 啟動WebServer   :8888
try:
    httpd.serve_forever()  # 等待用戶使用 WebServer
except:
    print("Closing the server.")
    httpd.server_close()  # 關閉 WebServer

    for x in range(len(a)):
        sheet_record.cell(row=x + 1, column=1).value = a[x]
        sheet_record.cell(row=x+1,column=2).value=b[x]
    wb_record.save("test_line.xlsx")
"""
{
{'destination': 'U5772b35df20e1a0ad36a4bc465916c51', 
'events': [
{'type': 'message', 'message': 
{'type': 'text', 'id': '16027358667156', 'text': '.'}, 
'webhookEventId': '01G26TNA5B24ZV3CAEQXF5615A', 'deliveryContext': 
{'isRedelivery': False}, 
'timestamp': 1651644212904, 
'source': {'type': 'user', 'userId': 'Ub47b1de232c030eb9f12505f3a015986'}, 
'replyToken': '93257297df16433aaa06263ffb916ab2', 'mode': 'active'}]}


"""